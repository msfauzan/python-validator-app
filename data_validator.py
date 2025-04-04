import pandas as pd
from difflib import SequenceMatcher
from openpyxl.styles import PatternFill
from openpyxl.comments import Comment
import re
import os
import db_utils

# Constants untuk nama kolom
COL_NAMA_PENERIMA = "nama_penerima"
COL_KATEGORI_PENERIMA = "kategori_penerima"
COL_NAMA_PEMBAYAR = "nama_pembayar"
COL_KATEGORI_PEMBAYAR = "kategori_pembayar"
COL_KODE_BANK = "cKdBank"
COL_STATUS_PENERIMA = "status_penerima"
COL_STATUS_PEMBAYAR = "status_pembayar"
COL_STT = "stt"
COL_TAHUN = "tahun"
COL_BULAN = "bulan"

class DataValidator:
    def __init__(self, fuzzy_match_threshold=0.9):
        self.reference_mapping = {
            "penerima": db_utils.get_mapping_data("ref_mapping_penerima"),
            "pembayar": db_utils.get_mapping_data("ref_mapping_pembayar"),
        }
        self.bank_codes = db_utils.get_bank_codes()
        self.fuzzy_match_threshold = fuzzy_match_threshold
        
        # Load STT category exceptions from config
        config = db_utils.load_config()
        self.stt_category_exceptions = config.get("validation", {}).get("stt_category_exceptions", {})
        self.status_mapping = db_utils.get_status_mapping()
        
        # Tambahkan prioritas kategori
        self.category_priority = ["B0", "C0", "F1", "F2"]

    def reload_reference_data(self):
        """Reload mapping dan bank codes dari database."""
        self.reference_mapping = {
            "penerima": db_utils.get_mapping_data("ref_mapping_penerima"),
            "pembayar": db_utils.get_mapping_data("ref_mapping_pembayar"),
        }
        self.bank_codes = db_utils.get_bank_codes()
        self.status_mapping = db_utils.get_status_mapping()

    def get_suggested_status(self, name):
        """Get suggested status based on keywords in the name."""
        name = str(name).upper()
        
        # Daftar kata-kata yang bisa diabaikan jika ada keyword negara
        company_words = {"PT", "PERSERO", "TBK", "LTD", "CV", "KOPERASI"}
        
        # Daftar kata pendek yang harus persis berdiri sendiri
        short_keywords = {"PT", "WHO", "UN", "TBK", "CV", "LTD", "INC", "CORP", "CO", "LLC", "PTE", "PVT", "BV", "NV", "SA", "GMBH", "AG", "SL", "SRL", "SAS", "SARL", "SPA", "SRL", "SNC", "SCS", "SCA", "SAR", "SASU", "SARLU"}

        # Jika keyword pendek ditemukan sekadar sebagai substring, abaikan
        for kw in short_keywords:
            if kw in name and not self.is_standalone_word(kw, name):
                # Jika ternyata hanya substring, kita lanjut tanpa men-flag
                return []

        # Cek dulu keyword negara (prioritas tertinggi)
        country_statuses = set()
        for keyword, statuses in self.status_mapping.items():
            if self.is_standalone_word(keyword, name):
                for status in statuses:
                    # Jika menemukan status negara
                    if status not in ["ID", "N1"]:
                        country_statuses.add(status)

        # Jika ada status negara, langsung return tanpa mengecek ID/N1
        if country_statuses:
            return list(country_statuses)
            
        # Jika tidak ada status negara, cek untuk ID/N1
        id_statuses = set()
        # Hanya cek kata-kata yang menunjukkan ID/N1 jika tidak ada status negara
        for keyword in company_words:
            if self.is_standalone_word(keyword, name):
                id_statuses.update(["ID", "N1"])
                break

        return list(id_statuses) if id_statuses else []

    def validate_bank_code(self, bank_name, bank_code):
        """
        Validasi kesesuaian nama bank dengan sandi bank

        Args:
            bank_name (str): Nama bank.
            bank_code (str): Kode bank.

        Returns:
            bool: True jika sesuai, False jika tidak.
        """
        if pd.isna(bank_code) or str(bank_code).strip() == '':
            return False

        try:
            bank_code = str(bank_code).zfill(3)
        except (ValueError, AttributeError):
            return False

        if bank_code not in self.bank_codes:
            return False

        reference_bank_name = self.bank_codes[bank_code].upper()
        bank_name = str(bank_name).upper() if not pd.isna(bank_name) else ""

        def clean_bank_name(name):
            name = str(name).upper()
            name = re.sub(r"[^\w\s]", " ", name)
            name = " ".join(name.split())
            # Hapus kata-kata lokasi/negara
            location_words = ["HONG KONG", "SINGAPORE", "INDONESIA", "CHINA", "JAPAN", ""]
            for loc in location_words:
                name = name.replace(loc, "")
            # Hapus kata-kata umum
            common_words = ["PT", "BANK", "PERSERO", "TBK", "LIMITED", "LTD"]
            for word in common_words:
                name = name.replace(word, "")
            return name.strip()

        clean_bank = clean_bank_name(bank_name)
        clean_reference = clean_bank_name(reference_bank_name)

        return clean_reference in clean_bank or clean_bank in clean_reference

    def is_same_bank(self, bank1, bank2):
        """
        Memeriksa apakah dua nama bank merujuk ke bank yang sama

        Args:
            bank1 (str): Nama bank pertama.
            bank2 (str): Nama bank kedua.

        Returns:
            bool: True jika sama, False jika tidak.
        """
        if pd.isna(bank1) or pd.isna(bank2) or str(bank1).strip() == '' or str(bank2).strip() == '':
            return False

        try:
            clean_bank1 = self.clean_bank_name(str(bank1))
            clean_bank2 = self.clean_bank_name(str(bank2))
        except (ValueError, AttributeError):
            return False

        def clean_bank_name(name):
            name = str(name).upper()
            name = re.sub(r"[^\w\s]", " ", name)
            name = " ".join(name.split())
            common_words = [
                "PT",
                "BANK",
                "PERSERO",
                "(PERSERO)",
                "TBK",
                "INCORPORATION",
                "CORPORATION",
                "LTD",
                "LIMITED",
                "INCORPORATED", 
            ]
            for word in common_words:
                name = name.replace(f" {word} ", " ")
                if name.startswith(f"{word} "):
                    name = name[len(word) :].strip()
                if name.endswith(f" {word}"):
                    name = name[: -len(word)].strip()
            return name

        clean_bank1 = clean_bank_name(bank1)
        clean_bank2 = clean_bank_name(bank2)

        if clean_bank1 == clean_bank2:
            return True

        def extract_bank_code(name):
            words = name.split()
            bank_codes = [
                word for word in words if len(word) in [2, 3, 4] and word.isalpha()
            ]
            return bank_codes[0] if bank_codes else None

        code1 = extract_bank_code(clean_bank1)
        code2 = extract_bank_code(clean_bank2)

        if code1 and code2 and code1 == code2:
            return True

        if (
            (clean_bank1 in clean_bank2 or clean_bank2 in clean_bank1)
            and (len(clean_bank1) == 0 or len(clean_bank2) == 0)
        ):
            return True

        return (
            SequenceMatcher(None, clean_bank1, clean_bank2).ratio()
            > self.fuzzy_match_threshold
        )

    def is_n1_category(self, stt_value):
        """
        Memeriksa apakah nilai stt termasuk dalam kategori N1.

        Args:
            stt_value (str): Nilai stt.

        Returns:
            bool: True jika termasuk kategori N1, False jika tidak.
        """
        return str(stt_value) in ["1NNN", "1000", "1901", "1902", "1903", "1904", "1905", "1911", "1912", "1906", "1907", "2NNN", "2000", "2901", "2902", "2903", "2904", "2905", "2911", "2912", "2906", "2907"]

    def is_category_allowed_for_stt(self, stt, category):
        """
        Check if a category is allowed for a specific STT code.
        
        Args:
            stt (str): The STT code
            category (str): The category to check
            
        Returns:
            bool: True if the category is allowed for this STT, False otherwise
        """
        allowed_categories = self.stt_category_exceptions.get(str(stt), [])
        return category in allowed_categories

    def is_standalone_word(self, word, text):
        """
        Memeriksa apakah sebuah kata berdiri sendiri dalam teks.
        Mendukung pengecekan multi-word keyword dan kasus khusus untuk singkatan.
        
        Args:
            word (str): Kata yang dicari
            text (str): Teks yang akan diperiksa
            
        Returns:
            bool: True jika kata berdiri sendiri, False jika bagian dari kata lain
        """
        if not text or not word:
            return False
            
        # Ubah ke uppercase untuk konsistensi
        word = word.upper()
        text = text.upper()

        # Daftar kata pendek yang harus persis berdiri sendiri
        short_keywords = {
            "PT", "WHO", "UN", "TBK", "CV", "LTD", "INC", "CORP", "CO", 
            "LLC", "PTE", "PVT", "BV", "NV", "SA", "GMBH", "AG", "SL", 
            "SRL", "SAS", "SARL", "SPA", "SNC", "SCS", "SCA", "SAR", 
            "SASU", "SARLU"
        }
        
        # Jika keyword adalah kata pendek, lakukan pengecekan lebih ketat
        if word in short_keywords:
            # Bersihkan teks dari karakter khusus kecuali underscore
            text = re.sub(r'[^\w\s]', ' ', text)
            # Normalisasi spasi menjadi single space
            text = ' '.join(text.split())
            # Tambahkan spasi di awal dan akhir untuk memudahkan pengecekan
            text = f" {text} "
            
            # Cari semua kemunculan kata dengan word boundaries
            matches = re.finditer(fr'\b{re.escape(word)}\b', text)
            
            for match in matches:
                # Ambil karakter sebelum dan sesudah match
                start, end = match.span()
                # Periksa apakah kata benar-benar berdiri sendiri
                if text[start-1] == " " and text[end] == " ":
                    return True
            return False
            
        # Untuk kata-kata normal atau multi-word keywords
        else:
            # Bersihkan teks dari karakter khusus
            text = re.sub(r'[^\w\s]', ' ', text)
            # Normalisasi spasi
            text = ' '.join(text.split())
            # Tambahkan spasi di awal dan akhir
            text = f" {text} "
            # Cari kata dengan spasi di sekitarnya
            return f" {word} " in text

    def check_specific_category(self, name, mapping_dict):
        """
        Memeriksa kategori spesifik berdasarkan keyword, dengan mempertimbangkan prioritas.
        Mengembalikan kategori yang ditemukan atau None jika tidak ada yang cocok.
        """
        name = str(name).upper()
        found_categories = {}
        
        # Cek semua keyword yang cocok terlebih dahulu
        for keyword, category in mapping_dict.items():
            if keyword.upper() in ["PT", "CV", "TBK"]:  # Skip generic company identifiers
                continue
            if self.is_standalone_word(keyword, name):
                if category not in found_categories:
                    found_categories[category] = keyword

        # Jika ada kategori yang ditemukan, prioritaskan berdasarkan self.category_priority
        if found_categories:
            for priority_category in self.category_priority:
                if priority_category in found_categories:
                    return priority_category

            # Jika tidak ada dalam priority list, return kategori pertama yang ditemukan
            return next(iter(found_categories.keys()))
                    
        # Jika tidak ada kategori spesifik, cek identifier umum
        for keyword, category in mapping_dict.items():
            if keyword.upper() in ["PT", "CV", "TBK"]:
                if self.is_standalone_word(keyword, name):
                    return category
                    
        return None

    def get_bank_category(self, name, status, bank_code, is_valid_code):
        """
        Menentukan kategori bank berdasarkan prioritas:
        1. Status yang disarankan (jika ada)
        2. Status dari data mentah
        3. Validasi kode bank
        
        Args:
            name (str): Nama bank
            status (str): Status dari data
            bank_code (str): Kode bank
            is_valid_code (bool): Validitas kode bank
            
        Returns:
            tuple: (suggested_category, suggested_status)
        """
        # Cek dulu status yang disarankan
        suggested_statuses = self.get_suggested_status(name)
        
        # Jika ada saran status
        if suggested_statuses:
            if "ID" in suggested_statuses and is_valid_code:
                return "C1", "ID"
            for status in suggested_statuses:
                if status != "N1":  # Skip N1 untuk pengecekan bank
                    return "C2", status
        
        # # Jika tidak ada saran status, gunakan status data
        # if status == "ID":
        #     return "C1", status
        # elif status:
        #     return "C2", status
            
        # # Jika tidak ada status yang valid, cek kode bank
        # if is_valid_code:
        #     return "C9", None
        
        return None, None

    def validate_c2_category(self, name, status, bank_code, is_valid_code):
        """
        Validasi apakah kategori C2 sudah tepat.
        Kategori C2 hanya untuk bank luar negeri.
        
        Returns:
            str or None: Kategori yang disarankan jika C2 tidak tepat, None jika sudah tepat
        """
        # Cek apakah ini adalah bank
        if not self.is_standalone_word("BANK", name):
            return "E0"  # Jika bukan bank, sarankan E0
            
        # Cek status untuk konfirmasi bank luar negeri
        suggested_statuses = self.get_suggested_status(name)
        
        # Jika ada suggested status
        if suggested_statuses:
            if "ID" in suggested_statuses:
                return "C1"  # Jika bank dalam negeri, harusnya C1
            if any(s not in ["ID", "N1"] for s in suggested_statuses):
                return None  # Status luar negeri ditemukan, C2 sudah tepat
                
        # Jika tidak ada suggested status, cek status data
        if status == "ID":
            return "C1"  # Jika bank dalam negeri, harusnya C1
        elif status and status != "N1":
            return None  # Ada status luar negeri, C2 sudah tepat
            
        # Jika tidak ada informasi status yang jelas
        if is_valid_code:
            return "C9"  # Jika kode bank valid tapi tidak jelas statusnya
        
        return "E0"  # Default jika tidak ada informasi yang cukup

    def process_file(self, input_file):
        """
        Memproses file Excel dan melakukan validasi.

        Args:
            input_file (str): Path ke file Excel input.

        Returns:
            tuple: (output_file, error_count, validation_results)
                - output_file (str): Path ke file Excel output.
                - error_count (int): Jumlah error yang ditemukan.
                - validation_results (list): List hasil validasi.
        """
        self.reload_reference_data()  # Pastikan memuat ulang mapping setiap kali proses
        try:
            # Validasi file exists dan extension
            if not os.path.exists(input_file):
                raise FileNotFoundError("File tidak ditemukan")
            
            # Check if trying to validate an already validated file
            if "_validated" in input_file:
                raise ValueError(
                    "File ini merupakan hasil validasi.\n"
                    "Silakan pilih file asli (tanpa suffix '_validated')"
                )

            df = pd.read_excel(input_file)
            
            # Get year and month from dataframe
            if "tahun" not in df.columns or "bulan" not in df.columns:
                raise ValueError(f"File Excel harus memiliki kolom tahun dan bulan")
            
            # Get first non-null values for year and month
            tahun = df["tahun"].dropna().iloc[0] if not df["tahun"].isna().all() else ""
            bulan = df["bulan"].dropna().iloc[0] if not df["bulan"].isna().all() else ""
            
            # Validate year and month
            try:
                tahun = int(tahun)
                bulan = int(bulan)
                if not (2000 <= tahun <= 2100) or not (1 <= bulan <= 12):
                    raise ValueError
            except (ValueError, TypeError):
                raise ValueError("Nilai tahun atau bulan tidak valid")

            # Generate dan buat folder output
            output_parent_folder = "Output"
            os.makedirs(output_parent_folder, exist_ok=True)
            output_folder_name = os.path.join(
                output_parent_folder,
                f"{os.path.splitext(os.path.basename(input_file))[0]}_{tahun}_{str(bulan).zfill(2)}_validated"
            )
            os.makedirs(output_folder_name, exist_ok=True)
            output_file = os.path.join(
                output_folder_name,
                os.path.basename(os.path.splitext(input_file)[0]) + f"_{tahun}_{str(bulan).zfill(2)}_validated.xlsx"
            )
            
            # Check if output file is currently open
            try:
                with open(output_file, 'a+b') as f:
                    pass
            except PermissionError:
                raise PermissionError(
                    f"File hasil validasi '{os.path.basename(output_file)}' sedang terbuka.\n"
                    "Silakan tutup file tersebut terlebih dahulu."
                )

            if not input_file.lower().endswith(('.xls', '.xlsx')):
                raise ValueError("Format file harus Excel (.xls atau .xlsx)")

            df = pd.read_excel(input_file)
            
            # Validasi minimal rows
            if len(df) == 0:
                raise ValueError("File Excel kosong")
                
            required_columns = [
                "nama_penerima",
                "kategori_penerima",
                "nama_pembayar",
                "kategori_pembayar",
                "stt",
            ]
            if not all(col in df.columns for col in required_columns):
                raise Exception(
                    f"File Excel tidak memiliki kolom yang dibutuhkan: {', '.join(required_columns)}"
                )

            output_df = df.copy()
            validation_results = []

            for idx, row in df.iterrows():
                # Check N1 first, if true then skip all other checks
                if self.is_n1_category(row.get("stt", "")):
                    # If N1, force both categories and statuses to N1
                    if row["kategori_penerima"] != "N1":
                        validation_results.append({
                            "row": idx + 2,
                            "column": "kategori_penerima",
                            "current": row["kategori_penerima"],
                            "suggested": "N1",
                            "name": row["nama_penerima"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": "N1",
                        })
                    if row["kategori_pembayar"] != "N1":
                        validation_results.append({
                            "row": idx + 2,
                            "column": "kategori_pembayar",
                            "current": row["kategori_pembayar"],
                            "suggested": "N1",
                            "name": row["nama_pembayar"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": "N1",
                        })
                    continue  # Skip all other validation checks for this row

                stt_value = row.get("stt", "")
                forced_categories = self.stt_category_exceptions.get(str(stt_value), [])

                if forced_categories:
                    # Ambil kategori pertama dari daftar
                    forced_category = forced_categories[0]
                    if row["kategori_penerima"] != forced_category:
                        validation_results.append({
                            "row": idx + 2,
                            "column": "kategori_penerima",
                            "current": row["kategori_penerima"],
                            "suggested": forced_category,
                            "name": row["nama_penerima"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": row.get("status_penerima", "")
                        })

                is_penerima_bank = (
                    "BANK" in str(row[COL_NAMA_PENERIMA]).upper() 
                    and row[COL_KATEGORI_PENERIMA] not in ["F1", "C0"]
                )
                is_pembayar_bank = (
                    "BANK" in str(row[COL_NAMA_PEMBAYAR]).upper() 
                    and row[COL_KATEGORI_PEMBAYAR] not in ["F1", "C0"]
                )
                is_valid_bank_code_penerima = self.validate_bank_code(
                    row[COL_NAMA_PENERIMA], row.get(COL_KODE_BANK, "")
                )
                is_valid_bank_code_pembayar = self.validate_bank_code(
                    row[COL_NAMA_PEMBAYAR], row.get(COL_KODE_BANK, "")
                )
                penerima_status = str(row.get(COL_STATUS_PENERIMA, "")).upper()
                pembayar_status = str(row.get(COL_STATUS_PEMBAYAR, "")).upper()
                is_same_bank = self.is_same_bank(row[COL_NAMA_PENERIMA], row[COL_NAMA_PEMBAYAR])

                suggested_category_penerima = None
                suggested_category_pembayar = None

                if suggested_category_pembayar is None:
                    # Add check for STT category exceptions before the main validation logic
                    if self.is_category_allowed_for_stt(row.get("stt", ""), row.get("kategori_penerima", "")):
                        # Skip validation for kategori_penerima if it's allowed for this STT
                        suggested_category_penerima = None
                        suggested_category_pembayar = row.get("kategori_pembayar", "")  # Keep original validation for pembayar
                    else:
                        if (
                            row["nama_penerima"] == row["nama_pembayar"]
                            and penerima_status == pembayar_status
                        ):
                            suggested_category_pembayar = "I0"
                        else:
                            for central_bank_keyword in self.reference_mapping["penerima"]:
                                if central_bank_keyword.upper() == "C0":
                                    continue
                                if (
                                    central_bank_keyword.upper()
                                    in str(row["nama_penerima"]).upper()
                                    and self.reference_mapping["penerima"][
                                        central_bank_keyword
                                    ]
                                    == "C0"
                                ):
                                    suggested_category_penerima = "C0"
                                    is_penerima_bank = False
                                    break

                            for central_bank_keyword in self.reference_mapping["pembayar"]:
                                if central_bank_keyword.upper() == "C0":
                                    continue
                                if (
                                    central_bank_keyword.upper()
                                    in str(row["nama_pembayar"]).upper()
                                    and self.reference_mapping["pembayar"][
                                        central_bank_keyword
                                    ]
                                    == "C0"
                                ):
                                    suggested_category_pembayar = "C0"
                                    is_pembayar_bank = False
                                    break

                            if suggested_category_penerima is None:
                                for q1_bank_keyword in self.reference_mapping["penerima"]:
                                    if q1_bank_keyword.upper() == "F1":
                                        continue
                                    if (
                                        q1_bank_keyword.upper()
                                        in str(row["nama_penerima"]).upper()
                                        and self.reference_mapping["penerima"][
                                            q1_bank_keyword
                                        ]
                                        == "F1"
                                    ):
                                        suggested_category_penerima = "F1"
                                        break

                            if suggested_category_pembayar is None:
                                for q1_bank_keyword in self.reference_mapping["pembayar"]:
                                    if q1_bank_keyword.upper() == "F1":
                                        continue
                                    if (
                                        q1_bank_keyword.upper()
                                        in str(row["nama_pembayar"]).upper()
                                        and self.reference_mapping["pembayar"][
                                            q1_bank_keyword
                                        ]
                                        == "F1"
                                    ):
                                        suggested_category_pembayar = "F1"
                                        break

                            if suggested_category_penerima is None and is_penerima_bank:
                                suggested_category_penerima, suggested_status_penerima = self.get_bank_category(
                                    row["nama_penerima"],
                                    row.get("status_penerima", ""),
                                    row.get("cKdBank", ""),
                                    is_valid_bank_code_penerima
                                )
                                
                                if suggested_category_penerima:
                                    # Jika ini bank penerima dengan kategori C1 atau C2,
                                    # dan pembayar adalah bank yang sama
                                    if (suggested_category_penerima in ["C1", "C2"] and 
                                        self.is_same_bank(row["nama_penerima"], row["nama_pembayar"])):
                                        # Sarankan kategori pembayar berdasarkan kategori penerima
                                        if suggested_category_penerima == "C1":
                                            suggested_category_pembayar = "C2"
                                        else:
                                            suggested_category_pembayar = "C1"

                            if suggested_category_pembayar is None and is_pembayar_bank:
                                suggested_category_pembayar, suggested_status_pembayar = self.get_bank_category(
                                    row["nama_pembayar"],
                                    row.get("status_pembayar", ""),
                                    row.get("cKdBank", ""),
                                    is_valid_bank_code_pembayar
                                )

                        if suggested_category_penerima is None and not is_penerima_bank:
                            suggested_category_penerima = self.check_specific_category(
                                row["nama_penerima"], 
                                self.reference_mapping["penerima"]
                            )

                        if suggested_category_pembayar is None and not is_pembayar_bank:
                            suggested_category_pembayar = self.check_specific_category(
                                row["nama_pembayar"], 
                                self.reference_mapping["pembayar"]
                            )

                if (
                    suggested_category_penerima
                    and row["kategori_penerima"] != suggested_category_penerima
                ):
                    validation_results.append(
                        {
                            "row": idx + 2,
                            "column": "kategori_penerima",
                            "current": row["kategori_penerima"],
                            "suggested": suggested_category_penerima,
                            "name": row["nama_penerima"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": penerima_status,
                        }
                    )

                if (
                    suggested_category_pembayar
                    and row["kategori_pembayar"] != suggested_category_pembayar
                ):
                    validation_results.append(
                        {
                            "row": idx + 2,
                            "column": "kategori_pembayar",
                            "current": row["kategori_pembayar"],
                            "suggested": suggested_category_pembayar,
                            "name": row["nama_pembayar"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": pembayar_status,
                        }
                    )

                # Tambahkan validasi khusus untuk kategori C2
                if row["kategori_pembayar"] == "C2":
                    suggested_category = self.validate_c2_category(
                        row["nama_pembayar"],
                        row.get("status_pembayar", ""),
                        row.get("cKdBank", ""),
                        is_valid_bank_code_pembayar
                    )
                    if suggested_category:
                        validation_results.append({
                            "row": idx + 2,
                            "column": "kategori_pembayar",
                            "current": "C2",
                            "suggested": suggested_category,
                            "name": row["nama_pembayar"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": row.get("status_pembayar", ""),
                        })

                # Check status penerima
                suggested_status_penerima = self.get_suggested_status(row["nama_penerima"])
                current_status_penerima = str(row.get("status_penerima", "")).upper()
                
                if "LTD" not in str(row["nama_penerima"]).upper():
                    if suggested_status_penerima and current_status_penerima not in suggested_status_penerima:
                        validation_results.append({
                            "row": idx + 2,
                            "column": "status_penerima",
                            "current": current_status_penerima,
                            "suggested": " or ".join(suggested_status_penerima),
                            "name": row["nama_penerima"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": current_status_penerima
                        })

                # Check status pembayar
                suggested_status_pembayar = self.get_suggested_status(row["nama_pembayar"])
                current_status_pembayar = str(row.get("status_pembayar", "")).upper()
                
                if "LTD" not in str(row["nama_pembayar"]).upper():
                    if suggested_status_pembayar and current_status_pembayar not in suggested_status_pembayar:
                        validation_results.append({
                            "row": idx + 2,
                            "column": "status_pembayar",
                            "current": current_status_pembayar,
                            "suggested": " or ".join(suggested_status_pembayar),
                            "name": row["nama_pembayar"],
                            "bank_code": row.get("cKdBank", ""),
                            "status": current_status_pembayar
                        })

            writer = pd.ExcelWriter(output_file, engine="openpyxl")
            output_df.to_excel(writer, index=False)

            worksheet = writer.sheets["Sheet1"]
            red_fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )

            for result in validation_results:
                col_idx = df.columns.get_loc(result["column"]) + 1
                cell = worksheet.cell(row=result["row"], column=col_idx)
                cell.fill = red_fill

                comment_text = f"Suggested category: {result['suggested']}\n"
                comment_text += f"Name: {result['name']}\n"
                if "bank_code" in result:
                    comment_text += f"Bank code: {result['bank_code']}\n"
                comment_text += f"Status: {result['status']}"

                cell.comment = Comment(comment_text, "Validator")

            # Buat kamus penggantian header
            rename_map = {
                "cKdBank": "cKdBank",
                "baris": "baris",
                "sandi_bank": "Sandi Bank",
                "tahun": "Thn",
                "bulan": "Bln",
                "tanggal": "Tgl",
                "nomer_identifikasi": "No. Identifikasi",
                "rekening": "Rek",
                "status_penerima": "SPn",
                "kategori_penerima": "KPn",
                "status_pembayar": "SPb",
                "kategori_pembayar": "KPb",
                "hubungan_keuangan": "HK",
                "sandi_negara": "NDK",
                "sandi_valuta": "Valuta",
                "nilai transaksi": "Nilai Transaksi",
                "stt": "STT",
                "nama_penerima": "Pelaku Penerima",
                "jenis_id_penerima": "Jns Id Pn",
                "nomor_id_penerima": "No Id Pn",
                "nama_pembayar": "Pelaku Pembayar",
                "jenis_id_pembayar": "Jns Id Pb",
                "nomor_id_pembayar": "No Id Pb",
                "bank_pengirim": "Bank Pengirim",
                "bank_penerima": "Bank Penerima",
                "detil_transaksi": "Keterangan Detail Transaksi",
                "info_DP": "info DP",
            }

            # Ubah header di worksheet utama
            for col_idx in range(1, worksheet.max_column + 1):
                old_header = worksheet.cell(row=1, column=col_idx).value
                new_header = rename_map.get(old_header, old_header)
                worksheet.cell(row=1, column=col_idx).value = new_header

            writer.close()

            # Mulai pemecahan file per cKdBank
            unique_banks = output_df['cKdBank'].dropna().unique()
            for bank_code in unique_banks:
                subset_df = output_df[output_df['cKdBank'] == bank_code].copy()
                if subset_df.empty:
                    continue

                split_file = os.path.join(
                    output_folder_name,
                    os.path.basename(os.path.splitext(input_file)[0]) + f"_{tahun}_{str(bulan).zfill(2)}_{bank_code}_validated.xlsx"
                )
                split_writer = pd.ExcelWriter(split_file, engine="openpyxl")
                subset_df.to_excel(split_writer, index=False)
                split_ws = split_writer.sheets["Sheet1"]

                for res in validation_results:
                    if res['bank_code'] == bank_code:
                        # Cari baris di subset_df yang sesuai
                        original_idx = res['row'] - 2  # 0-based index
                        if original_idx in subset_df.index:
                            # Dapatkan baris 'baru' di subset
                            new_row = subset_df.index.get_loc(original_idx) + 2
                            new_col = subset_df.columns.get_loc(res["column"]) + 1
                            split_cell = split_ws.cell(row=new_row, column=new_col)
                            split_cell.fill = red_fill
                            comment_text = (
                                f"Suggested category: {res['suggested']}\n"
                                f"Name: {res['name']}\n"
                                f"Bank code: {res.get('bank_code','')}\n"
                                f"Status: {res['status']}"
                            )
                            split_cell.comment = Comment(comment_text, "Validator")
                
                # Ubah header di worksheet split
                for col_idx in range(1, split_ws.max_column + 1):
                    old_header = split_ws.cell(row=1, column=col_idx).value
                    new_header = rename_map.get(old_header, old_header)
                    split_ws.cell(row=1, column=col_idx).value = new_header

                split_writer.close()

            return output_file, len(validation_results), validation_results

        except Exception as e:
            raise Exception(f"Error processing file: {str(e)}")